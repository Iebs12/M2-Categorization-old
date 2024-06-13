# Import necessary libraries
from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
import os
import qdrant_client
from dotenv import load_dotenv


# Import llama_index components

from llama_index.core import VectorStoreIndex, get_response_synthesizer
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.retrievers import VectorIndexRetriever
from llama_index.core import Document
from llama_index.vector_stores.qdrant import QdrantVectorStore
from llama_index.embeddings.openai import OpenAIEmbedding
from llama_index.core.ingestion import IngestionPipeline
from llama_index.core.response.pprint_utils import pprint_response
from llama_index.core.query_engine import CustomQueryEngine
from llama_index.core.retrievers import BaseRetriever
from llama_index.core.response_synthesizers import BaseSynthesizer
from llama_index.llms.openai import OpenAI
from llama_index.core import PromptTemplate

# Initialize Flask application
app = Flask(__name__)

def extractor(file_path):
    """
    Extracts data from an Excel file.
    
    Args:
        file_path (str): The path to the Excel file.

    Returns:
        list: A list of dictionaries containing extracted data.
    """
    # Load the workbook in read-only mode
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active  # Select the first sheet

    data = []  # List to store row data as dictionaries

    # Get the header row
    header = [cell for cell in next(ws.iter_rows(values_only=True))]
    col_index = {name: index for index, name in enumerate(header)}  # Column name to index mapping

    # Iterate over the rows starting from the second row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not all(cell is None for cell in row):  # Skip rows with all None values
            row_dict = {
                'title': row[col_index['Title']],
                'abstract': row[col_index['Abstract']],
                'description': row[col_index['English description']],
                'claims': row[col_index['Claims']]
            }
            data.append(row_dict)

    return data

def newFileSaver(relevancy, file_path):
    """
    Saves relevancy data to a new column in the Excel file.
    
    Args:
        relevancy (list): A list of tuples containing relevancy status and comments.
        file_path (str): The path to the Excel file.

    Returns:
        str: The file path where the updated file is saved.
    """
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    relevancy_header = 'Relevancy predicted'
    comments_header = 'Comments made'

    empty_column = None
    for cell in sheet[1]:
        if cell.value is None:
            empty_column = cell.column
            break

    if empty_column is None:
        empty_column = sheet.max_column + 1

    # Add headers to the new columns
    sheet.cell(row=1, column=empty_column, value=relevancy_header)
    sheet.cell(row=1, column=empty_column + 1, value=comments_header)

    # Add relevancy data to the new columns
    for i, (status, comment) in enumerate(relevancy, start=2):
        sheet.cell(row=i, column=empty_column, value=status)
        sheet.cell(row=i, column=empty_column + 1, value=comment)

    workbook.save(filename=file_path)
    
    # Create a new workbook for entries with 'R'
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # Copy headers to the new workbook
    for col_num, cell in enumerate(sheet[1], 1):
        new_sheet.cell(row=1, column=col_num, value=cell.value)
    
    # Filter rows with 'R' and copy to new workbook
    new_row_idx = 2
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[empty_column-1].value == 'R':
            for col_num, cell in enumerate(row, 1):
                new_sheet.cell(row=new_row_idx, column=col_num, value=cell.value)
            new_row_idx += 1

    # Save the new workbook
    new_file_path = os.path.splitext(file_path)[0] + '_filtered.xlsx'
    new_workbook.save(filename=new_file_path)

    return file_path, new_file_path

def extract_reason(text):
    """
    Extracts reason text from a given string.
    
    Args:
        text (str): The input text.

    Returns:
        str: Extracted reason text.
    """
    parts = text.split("Reason: ", 1)
    return parts[1] if len(parts) > 1 else ""

def extract_related(text):
    """
    Checks if the text contains the string '1R1'.
    
    Args:
        text (str): The input text.

    Returns:
        bool: True if '1R1' is found, False otherwise.
    """
    return '1R1' in text

def backend(dict_item, user_query):
    """
    Backend function to process a document and a query.
    
    Args:
        dict_item (dict): A dictionary representing the document.
        user_query (str): The user's query.

    Returns:
        tuple: A tuple containing related status and reason.
    """
    load_dotenv()  # Load environment variables

    llama_api_key = os.getenv('LLAMA_CLOUD_API_KEY')
    if llama_api_key is None:
        raise ValueError("LLAMA_CLOUD_API_KEY not found in environment variables")

    # Create documents from the dictionary item
    documents = [Document(text=f"{key}: {val}") for key, val in dict_item.items()]

    client = qdrant_client.QdrantClient(location=":memory:")
    vector_store = QdrantVectorStore(client=client, collection_name="test_store")

    # Ingest documents into a vector store
    pipeline = IngestionPipeline(
        transformations=[
            SentenceSplitter(chunk_size=128, chunk_overlap=5),
            OpenAIEmbedding(),
        ],
        vector_store=vector_store,
    )
    pipeline.run(documents=documents)
    index = VectorStoreIndex.from_vector_store(vector_store)

    class RAGStringQueryEngine(CustomQueryEngine):
        """
        Custom Query Engine for RAG (Retrieval-Augmented Generation).
        """
        retriever: BaseRetriever
        response_synthesizer: BaseSynthesizer
        llm: OpenAI
        qa_prompt: PromptTemplate

        def custom_query(self, query_str: str):
            """
            Perform a custom query.
            
            Args:
                query_str (str): The query string.

            Returns:
                str: The response from the query.
            """
            nodes = self.retriever.retrieve(query_str)
            context_str = "\n\n".join([n.node.get_content() for n in nodes])

            # # Uncomment the following to print context string as well.
            # print('--'*50)
            # print(context_str)
            # print('--'*50)

            response = self.llm.complete(qa_prompt.format(context_str=context_str, query_str=query_str))
            return str(response)

    qa_prompt = PromptTemplate(
        "You are an AI assistant that predicts relevancy of a 'Document' with a certain 'Statement'. If it is Relevant then return output as '1R1', otherwise '0R0'. If output is '1R1', then state the 'Reason'  which makes it relevant with the help of information present in 'Document'. \n"
       "For example 1:\n"
             
       "Document:" + ''' title: Composition, application of the composition, cosmetic preparation hydrogel bio-mask in the form of a compress, method of manufacturing the preparation
       Background of the invention.
       hydrogel bio-mask composed of natural materials and active ingredients, designed for cosmetic applications to enhance skin health. The hydrogel matrix provides a natural and effective medium for delivering active ingredients to the skin. the composition of the hydrogel bio-mask and its natural active ingredients. The following are the key points regarding the specific ingredients mentioned
        Hydrogel Matrix: The document emphasizes the use of a hydrogel matrix obtained from natural sources. Natural Active Ingredients: The hydrogel bio-mask includes various natural active ingredients intended for cosmetic use.''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '0R0' \n"
       "Reason:  It is not mentioning the use of Mannuronic acid, alginate, or avocado but having skin claim for cosmetics \n"
             
       "For example 2:\n"
             
       "Document:" + ''' the use of mannuronic acid derivatives and alginate from algae in cosmetic formulations aimed at improving skin health by providing anti-photoaging benefits, moisture retention, antioxidant protection, and enzyme inhibition. The derivatives form an invisible film on the skin, protecting against UV damage and maintaining a moist environment. They exhibit strong antioxidant capabilities and inhibit enzymes like tyrosinase and elastase, reducing melanin production and collagen degradation.
        The primary focus of the patent is on alginate oligosaccharide derivatives derived from brown algae. These are used for their moisture absorption, antioxidation, and enzyme inhibition properties in skincare products. ''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '1R1' \n"
       "Reason :  Mannuronic acid and alginate from algae is used for different skin claims in a cosmetic product \n"
 
        "For example 3:\n"
 
        "Document:" + ''' title: Use of brown algae water extract for preparing blue light resistant skin external product
       Background of the invention.
       using brown algae extract containing fucoidan for preparing topical skin care products that protect against blue light exposure. These products aim to improve skin health by reducing wrinkles and enhancing brightness, particularly for individuals frequently exposed to blue light. The invention emphasizes the benefits of fucoidan in long-term skin care.
       The present invention provides a use of a brown algae extract for preparing a skin topical product for anti-blue light, wherein the product is provided to a subject exposed to blue light, and the brown algae extract contains fucoidan.''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '1R1' \n"
       "Reason: Alginate from Brown Algae is used for protecting against blue light in skincare products \n"
 
       "Using the below given Document and Statement , provide the Output and Reason"
        "Document: {context_str}\n"
        "Statement: {query_str}\n"
        "Output: "
        "Reason: "
    )

    retriever = VectorIndexRetriever(index=index, similarity_top_k=5)
    response_synthesizer = get_response_synthesizer(response_mode="tree_summarize")
    llm = OpenAI(model="gpt-3.5-turbo")

    query_engine = RAGStringQueryEngine(
        retriever=retriever,
        response_synthesizer=response_synthesizer,
        llm=llm,
        qa_prompt=qa_prompt,
    )
    response = query_engine.query(user_query)

    # Uncomment the following if you need to see the responses for each input on terminal.
    pprint_response(response)

    response_str = str(response)
    related = extract_related(response_str)
    reason = extract_reason(response_str)
    return related, reason

@app.route("/", methods=['GET', 'POST'])
def process_file():
    """
    Flask route to process the uploaded file and query.
    
    Returns:
        Response: JSON response containing the path to the updated file.
    """
    data = request.json
    query = data.get('query')
    file_path = data.get('file_path')
    datalist = extractor(file_path)

    if datalist[-1]['title'] is None:
        datalist.pop()

    # Initialize the relevancy list
    relevancy = []

    # Iterate over each dictionary in the datalist
    for dict_item in datalist:
        # Call the backend function with the current dictionary
        result = backend(dict_item, query)
        # Check the first element of the tuple and set 'R' or 'NR' accordingly
        status = "R" if result[0] else "NR"
        # Append the modified result to the relevancy list
        relevancy.append((status, result[1]))

    outputFilePath, newFilePath = newFileSaver(relevancy, file_path)
    
    return jsonify({'Path': outputFilePath, 'FilteredPath': newFilePath})

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)
