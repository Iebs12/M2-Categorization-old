from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
import os
from dotenv import load_dotenv

# Importing necessary modules from llama_index
from llama_index.core import VectorStoreIndex, get_response_synthesizer, Document, StorageContext
from llama_index.core.node_parser import SentenceSplitter
from llama_index.retrievers.bm25 import BM25Retriever
from llama_index.core.response.pprint_utils import pprint_response
from llama_index.core.query_engine import CustomQueryEngine
from llama_index.core.retrievers import BaseRetriever, VectorIndexRetriever, QueryFusionRetriever
from llama_index.core.response_synthesizers import BaseSynthesizer
from llama_index.llms.openai import OpenAI
from llama_index.core import PromptTemplate

# Initialize the Flask application
app = Flask(__name__)

def extractor(file_path):
    """
    Extracts data from an Excel file and returns it as a list of dictionaries.
    
    :param file_path: Path to the Excel file.
    :return: List of dictionaries with data extracted from the file.
    """
    # Load the workbook in read-only mode
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active  # Select the first sheet

    data = []  # Initialize an empty list to store extracted data

    # Get the header row
    header = [cell for cell in next(ws.iter_rows(values_only=True))]
    col_index = {name: index for index, name in enumerate(header)}  # Map column names to indices

    # Iterate over rows, starting from the second row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not all(cell is None for cell in row):  # Skip empty rows
            row_dict = {
                'title': row[col_index['Title']],
                'abstract': row[col_index['Abstract']],
                'description': row[col_index['English description']],
                'claims': row[col_index['Claims']]
            }
            data.append(row_dict)  # Add the row data to the list

    return data

def newFileSaver(relevancy, file_path):
    """
    Saves relevancy data to a new column in the existing Excel file.
    
    :param relevancy: List of tuples with relevancy and comments.
    :param file_path: Path to the Excel file.
    :return: Updated file path.
    """
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active  # Select the active worksheet

    relevancy_header = 'Relevancy predicted'
    comments_header = 'Comments made'

    # Find an empty column
    empty_column = None
    for cell in sheet[1]:
        if cell.value is None:
            empty_column = cell.column
            break
    if empty_column is None:
        empty_column = sheet.max_column + 1

    # Add headers to the first row of the new columns
    sheet.cell(row=1, column=empty_column, value=relevancy_header)
    sheet.cell(row=1, column=empty_column + 1, value=comments_header)

    # Add the values from the relevancy list to the new columns
    for i, (relevancy, comment) in enumerate(relevancy, start=2):
        sheet.cell(row=i, column=empty_column, value=relevancy)
        sheet.cell(row=i, column=empty_column + 1, value=comment)

    # Save the updated workbook
    workbook.save(filename=file_path)
    
    # Create a new workbook for entries with 'R'
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # Copy headers to the new workbook
    for col_num, cell in enumerate(sheet[1], 1):
        new_sheet.cell(row=1, column=col_num, value=cell.value)
    
    # Filter rows with 'R' and copy to new workbook
    new_row_idx = 2
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[empty_column-1].value == 'R':
            for col_num, cell in enumerate(row, 1):
                new_sheet.cell(row=new_row_idx, column=col_num, value=cell.value)
            new_row_idx += 1

    # Save the new workbook
    new_file_path = os.path.splitext(file_path)[0] + '_filtered.xlsx'
    new_workbook.save(filename=new_file_path)

    return file_path, new_file_path

def extract_reason(text):
    """
    Extracts the reason from the response text.
    
    :param text: Text to extract the reason from.
    :return: Extracted reason.
    """
    parts = text.split("Reason: ", 1)
    return parts[1] if len(parts) > 1 else ""

def extract_related(text):
    """
    Checks if the text indicates a relation by looking for '1R1'.
    
    :param text: Text to check for relation.
    :return: True if related, otherwise False.
    """
    return '1R1' in text

def backend(dict_item, user_query):
    """
    Processes the given dictionary and user query to predict relevancy.
    
    :param dict_item: Dictionary containing document data.
    :param user_query: User query string.
    :return: Tuple with related status and reason.
    """
    load_dotenv()  # Load environment variables from .env file
    llama_api_key = os.getenv('LLAMA_CLOUD_API_KEY')
    if llama_api_key is None:
        raise ValueError("LLAMA_CLOUD_API_KEY not found in environment variables")

    # Create Document objects from dictionary items
    documents = [Document(text=f"{key}: {val}") for key, val in dict_item.items()]

    # Create and configure the VectorStoreIndex
    splitter = SentenceSplitter(chunk_size=128, chunk_overlap=5)
    index = VectorStoreIndex.from_documents(documents, transformations=[splitter])

    class RAGStringQueryEngine(CustomQueryEngine):
        """
        Custom Query Engine using RAG (Retrieval-Augmented Generation) approach.
        """
        retriever: BaseRetriever
        response_synthesizer: BaseSynthesizer
        llm: OpenAI
        qa_prompt: PromptTemplate

        def custom_query(self, query_str: str):
            """
            Custom query method to retrieve and synthesize response.
            
            :param query_str: Query string.
            :return: Response string.
            """
            nodes = self.retriever.retrieve(query_str)
            context_str = "\n\n".join([n.node.get_content() for n in nodes])

            # # Uncomment the following to print context string as well.
            # print('--'*50)
            # print(context_str)
            # print('--'*50)

            response = self.llm.complete(
                qa_prompt.format(context_str=context_str, query_str=query_str)
            )
            return str(response)

    # Configure retrievers
    retriever1 = VectorIndexRetriever(index=index, similarity_top_k=5)
    nodes = splitter.get_nodes_from_documents(documents)
    storage_context = StorageContext.from_defaults()
    storage_context.docstore.add_documents(nodes)
    retriever2 = BM25Retriever.from_defaults(nodes=nodes, similarity_top_k=5)

    retriever = QueryFusionRetriever(
        [retriever1, retriever2],
        similarity_top_k=5,
        num_queries=4,
        mode="reciprocal_rerank",
        use_async=True,
        verbose=True,
    )

    response_synthesizer1 = get_response_synthesizer(response_mode="tree_summarize")

    qa_prompt = PromptTemplate(
        "You are an AI assistant that predicts relevancy of a 'Document' with a certain 'Statement'. If it is even a little relevant then return output as '1R1', otherwise '0R0'. If output is '1R1', then state the 'Reason'  which makes it relevant with the help of information present in 'Document'. \n"
       "For example 1:\n"
             
       "Document:" + ''' title: Composition, application of the composition, cosmetic preparation hydrogel bio-mask in the form of a compress, method of manufacturing the preparation
       Background of the invention.
       hydrogel bio-mask composed of natural materials and active ingredients, designed for cosmetic applications to enhance skin health. The hydrogel matrix provides a natural and effective medium for delivering active ingredients to the skin. the composition of the hydrogel bio-mask and its natural active ingredients. The following are the key points regarding the specific ingredients mentioned
        Hydrogel Matrix: The document emphasizes the use of a hydrogel matrix obtained from natural sources. Natural Active Ingredients: The hydrogel bio-mask includes various natural active ingredients intended for cosmetic use.''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '0R0' \n"
       "Reason:  It is not mentioning the use of Mannuronic acid, alginate, or avocado but having skin claim for cosmetics \n"
             
       "For example 2:\n"
             
       "Document:" + ''' the use of mannuronic acid derivatives and alginate from algae in cosmetic formulations aimed at improving skin health by providing anti-photoaging benefits, moisture retention, antioxidant protection, and enzyme inhibition. The derivatives form an invisible film on the skin, protecting against UV damage and maintaining a moist environment. They exhibit strong antioxidant capabilities and inhibit enzymes like tyrosinase and elastase, reducing melanin production and collagen degradation.
        The primary focus of the patent is on alginate oligosaccharide derivatives derived from brown algae. These are used for their moisture absorption, antioxidation, and enzyme inhibition properties in skincare products. ''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '1R1' \n"
       "Reason :  Mannuronic acid and alginate from algae is used for different skin claims in a cosmetic product \n"
 
        "For example 3:\n"
 
        "Document:" + ''' title: Use of brown algae water extract for preparing blue light resistant skin external product
       Background of the invention.
       using brown algae extract containing fucoidan for preparing topical skin care products that protect against blue light exposure. These products aim to improve skin health by reducing wrinkles and enhancing brightness, particularly for individuals frequently exposed to blue light. The invention emphasizes the benefits of fucoidan in long-term skin care.
       The present invention provides a use of a brown algae extract for preparing a skin topical product for anti-blue light, wherein the product is provided to a subject exposed to blue light, and the brown algae extract contains fucoidan.''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '1R1' \n"
       "Reason: Alginate from Brown Algae is used for protecting against blue light in skincare products \n"
 
       "Using the below given Document and Statement , provide the Output and Reason"
        "Document: {context_str}\n"
        "Statement: {query_str}\n"
        "Output: "
        "Reason: "
    )

    llm = OpenAI(model="gpt-3.5-turbo")

    query_engine = RAGStringQueryEngine(
        retriever=retriever,
        response_synthesizer=response_synthesizer1,
        llm=llm,
        qa_prompt=qa_prompt,
    )

    response = query_engine.query(user_query)
    pprint_response(response, show_source=True)
    response_str = str(response)

    related = extract_related(response_str)
    reason = extract_reason(response_str)
    return related, reason

@app.route("/", methods=['GET', 'POST'])
def process_file():
    """
    Endpoint to process the uploaded file and user query.
    
    :return: JSON response with the path to the updated file.
    """
    data = request.json
    query = data.get('query')
    file_path = data.get('file_path')

    datalist = extractor(file_path)
    if datalist and datalist[-1]['title'] is None:
        datalist.pop()  # Remove the last item if it has no title

    relevancy = []  # Initialize the relevancy list

    # Process each dictionary item
    for dict_item in datalist:
        result = backend(dict_item, query)
        status = "R" if result[0] else "NR"
        relevancy.append((status, result[1]))

    outputFilePath, newFilePath = newFileSaver(relevancy, file_path)
    
    return jsonify({'Path': outputFilePath, 'FilteredPath': newFilePath})

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)
